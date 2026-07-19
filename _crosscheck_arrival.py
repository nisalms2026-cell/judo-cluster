# -*- coding: utf-8 -*-
"""Cross-check Arrival Plan.xlsx vs data/arrival.json (+ directory)."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
XLSX = BASE / "Arrival Plan.xlsx"
ARR = BASE / "data" / "arrival.json"
DIR = BASE / "data" / "directory.json"

# Map Excel team labels -> dashboard org keys
ORG_ALIASES = {
    "assam rifles": "ASSAM RIFLES",
    "himachal pradesh police": "HIMACHAL PRADESH POLICE",
    "himachal pradesh police (half team)": "HIMACHAL PRADESH POLICE",
    "border security force": "BSF",
    "odisha": "ORISSA POLICE",
    "orissa": "ORISSA POLICE",
    "itbp wushu team": "ITBP",
    "itbp": "ITBP",
    "ssb": "SSB",
    "haryana police": "HARYANA",
    "haryana": "HARYANA",
    "kerala police": "KERALA POLICE",
    "karnataka state police": "KARNATAKA POLICE",
    "karnataka police": "KARNATAKA POLICE",
    "chhattisgarh": "CHATTISGARH",
    "chattisgarh": "CHATTISGARH",
    "uttarakhand": "UTTARAKHAND POLICE",
    "gujarat": "Gujrat",
    "gujrat": "Gujrat",
    "bihar": "BIHAR POLICE",
    "uttar pradesh": "UTTARPRADESH",
    "uttarpradesh": "UTTARPRADESH",
    "j&k police": "JAMMU & KASHMIR",
    "j&k": "JAMMU & KASHMIR",
    "j&k (1)": "JAMMU & KASHMIR",
    "jammu & kashmir": "JAMMU & KASHMIR",
    "punjab": "PUNJAB POLICE",
    "punjab (16 a)": "PUNJAB POLICE",
    "crpf": "CRPF",
    "cisf": "CISF",
    "cisf (18 a)": "CISF",
    "delhi police": "DELHI POLICE",
    "arunachal pradesh": "ARUNACHAL PRADESH",
    "chandigarh police": "CHANDIGARH POLICE",
    "maharashtra police": "MAHARASHTRA POLICE",
    "rajasthan": "RAJASTHAN POLICE",
    "tamil nadu": "TAMILNADU",
    "west bengal police": "WEST BENGAL POLICE",
    "rpf": "RPF",
    "assam police": "ASSAM POLICE",
    "assam police (21 a)": "ASSAM POLICE",
    "jharkhand": "JHARKHAND",
    "madhya pradesh": "MADHYA PRADESH POLICE",
    "mizoram police": "MIZORAM POLICE",
    "manipur police": "MANIPUR POLICE",
    "sikkim": "SIKKIM POLICE",
    "sikkim (02)": "SIKKIM POLICE",
    "sikkim (sb)": "SIKKIM POLICE",
    "andhra pradesh": "Andhra Pradesh",
    "andhra police": "Andhra Pradesh",
    "andhra pradesh police": "Andhra Pradesh",
    "telangana police": "TELANGANA POLICE",
    "goa": "GOA",
    "nsg": "NSG",
    "ngo": "NSG",
}


def norm_team(name: str) -> str:
    s = " ".join(str(name or "").lower().split())
    s = re.sub(r"\(half team\)", "", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def map_org(team: str) -> str | None:
    key = norm_team(team)
    if key in ORG_ALIASES:
        return ORG_ALIASES[key]
    # strip parenthetical suffixes like (6 A)
    key2 = re.sub(r"\([^)]*\)", "", key).strip()
    if key2 in ORG_ALIASES:
        return ORG_ALIASES[key2]
    for alias, org in ORG_ALIASES.items():
        if key.startswith(alias) or alias.startswith(key):
            return org
    return None


def station_mode(station: str) -> tuple[str, str]:
    s = (station or "").lower()
    if "flight" in s or "airport" in s or "rgia" in s or "shamshabad" in s:
        return "flight", "Rajiv Gandhi Intl Airport (Shamshabad)"
    if "bus" in s:
        return "bus", "Bus Alighting Point"
    if "await" in s:
        return "", ""
    if "secunderabad" in s:
        return "rail", "Secunderabad Jn"
    if "charlapalli" in s or "charlapali" in s:
        return "rail", "Charlapalli Rly. Stn."
    if "kacheguda" in s or "kachiguda" in s:
        return "rail", "Kacheguda Station"
    if "hyderabad" in s:
        return "rail", "Hyderabad Rly. Station"
    return "rail", station or ""


def first_arrival_line(raw: str) -> str:
    text = str(raw or "").replace("\r", "\n")
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    return lines[0] if lines else ""


def phone_digits(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        v = int(v)
    return re.sub(r"\D", "", str(v))


def parse_excel():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    cur_station = ""
    out = []
    for r in range(2, ws.max_row + 1):
        sl = ws.cell(r, 1).value
        st = ws.cell(r, 2).value
        team = ws.cell(r, 4).value
        mgr = ws.cell(r, 5).value
        phone = ws.cell(r, 6).value
        arr = ws.cell(r, 7).value
        det = ws.cell(r, 8).value
        if st:
            cur_station = str(st).strip()
        # "Travel plan awaited" is in No. of Team column on this sheet
        team_no = ws.cell(r, 3).value
        if team_no and "await" in str(team_no).lower():
            cur_station = str(team_no).strip()
        if not team:
            continue
        team_s = " ".join(str(team).split())
        # section banners in Team Name column
        low = team_s.lower()
        if low.startswith("arrival") or low.startswith("travel"):
            cur_station = team_s
            continue
        mode, hub = station_mode(cur_station)
        awaited = "await" in cur_station.lower()
        out.append(
            {
                "excel_row": r,
                "sl": sl,
                "section": cur_station,
                "team": team_s,
                "org": map_org(team_s),
                "manager": " ".join(str(mgr).split()) if mgr else "",
                "phone": phone_digits(phone),
                "arrival_raw": str(arr).strip() if arr else "",
                "arrival": first_arrival_line(arr),
                "details": str(det).strip() if det else "",
                "mode": "" if awaited else mode,
                "station": "" if awaited else hub,
                "status": "awaited" if awaited else ("planned" if (mode == "bus" or arr or det) else "awaited"),
            }
        )
    return out


def main():
    excel_rows = parse_excel()
    arr = json.loads(ARR.read_text(encoding="utf-8"))
    directory = json.loads(DIR.read_text(encoding="utf-8"))
    by_org = {r["org"]: r for r in arr.get("rows") or []}
    dir_map = {r["org"]: (r.get("manager") or {}) for r in directory.get("rows") or []}

    print("=" * 72)
    print(f"Excel: {XLSX.name}  |  rows parsed: {len(excel_rows)}")
    print("=" * 72)

    unmatched = [x for x in excel_rows if not x["org"]]
    if unmatched:
        print("\nUNMAPPED Excel teams:")
        for x in unmatched:
            print(f"  row {x['excel_row']}: {x['team']}")

    # Collapse excel to primary row per org (first occurrence = primary hub)
    primary = {}
    extras = {}
    for x in excel_rows:
        org = x["org"]
        if not org:
            continue
        if org not in primary:
            primary[org] = x
        else:
            extras.setdefault(org, []).append(x)

    print("\n--- TRAVEL CROSS-CHECK (Excel primary vs dashboard) ---\n")
    mismatches = []
    missing_in_dash = []
    for org, x in sorted(primary.items()):
        dash = by_org.get(org)
        if not dash:
            missing_in_dash.append(org)
            print(f"[MISSING IN DASH] {org}  Excel: {x['station']} | {x['arrival']}")
            continue
        t = dash.get("travel") or {}
        issues = []
        if (t.get("status") or "") != x["status"]:
            issues.append(f"status {t.get('status')!r} -> {x['status']!r}")
        if x["status"] == "planned":
            if (t.get("mode") or "") != x["mode"]:
                issues.append(f"mode {t.get('mode')!r} -> {x['mode']!r}")
            # station soft-compare
            ds = (t.get("station") or "").lower()
            xs = (x["station"] or "").lower()
            if xs and xs not in ds and ds not in xs:
                # allow Charlapalli listed under Secunderabad section if details say so
                if "charlapalli" in (x["details"] or "").lower() and "charlapalli" in ds:
                    pass
                elif x["mode"] == "bus" and "bus" in ds:
                    pass
                else:
                    issues.append(f"station {t.get('station')!r} -> {x['station']!r}")
            # arrival: compare first stamp digits
            da = re.sub(r"\s+", "", t.get("arrival") or "")
            xa = re.sub(r"\s+", "", x["arrival"] or "")
            if xa and xa[:10] not in da and da[:10] not in xa:
                issues.append(f"arrival {t.get('arrival')!r} -> {x['arrival']!r}")
        if issues:
            mismatches.append((org, issues, x, t))
            print(f"[DIFF] {org}")
            for i in issues:
                print(f"       {i}")
            print(f"       Excel mgr/phone: {x['manager']} / {x['phone']}")
        else:
            print(f"[OK]   {org}")

    # Orgs in dashboard not mentioned in excel primary
    excel_orgs = set(primary)
    dash_only = sorted(set(by_org) - excel_orgs)
    if dash_only:
        print("\n--- In dashboard but not as Excel primary team ---")
        for org in dash_only:
            t = by_org[org].get("travel") or {}
            print(f"  {org}: {t.get('status')} {t.get('mode')} {t.get('station')} {t.get('arrival')}")

    print("\n--- DIRECTORY CONTACTS (Excel vs directory.json) ---\n")
    contact_diff = 0
    for org, x in sorted(primary.items()):
        if not x["phone"] and not x["manager"]:
            continue
        m = dir_map.get(org) or {}
        dphone = phone_digits(m.get("phone"))
        if x["phone"] and dphone != x["phone"]:
            contact_diff += 1
            print(f"[PHONE] {org}: dash {dphone or '—'} -> excel {x['phone']} ({x['manager']})")
        elif x["manager"] and (m.get("name") or "").lower() not in x["manager"].lower() and x["manager"].lower() not in (m.get("name") or "").lower():
            contact_diff += 1
            print(f"[NAME]  {org}: dash {m.get('name')!r} -> excel {x['manager']!r}")

    print("\n" + "=" * 72)
    print(f"Summary: Excel teams mapped={len(primary)}  travel diffs={len(mismatches)}  "
          f"missing in dash={len(missing_in_dash)}  contact diffs={contact_diff}  unmapped={len(unmatched)}")
    print("=" * 72)

    # Write machine-readable report
    report = {
        "excel_file": str(XLSX),
        "mapped": len(primary),
        "travel_diffs": [{"org": o, "issues": iss} for o, iss, _, _ in mismatches],
        "missing_in_dashboard": missing_in_dash,
        "unmapped_excel_teams": [x["team"] for x in unmatched],
        "contact_diffs": contact_diff,
    }
    (BASE / "_arrival_crosscheck_report.json").write_text(
        json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print("Wrote _arrival_crosscheck_report.json")


if __name__ == "__main__":
    main()

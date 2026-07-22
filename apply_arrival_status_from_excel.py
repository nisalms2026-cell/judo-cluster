"""Sync arrived_strength in data/arrival.json from TGPA_Arrival_Status.xlsx."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

import store
from org_names import canonical_org

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "TGPA_Arrival_Status.xlsx"

# Extra labels not covered by org_names.py
_ORG_ALIASES = {
    "UTTAR PRADESH": "UTTARPRADESH",
}


def map_org(label: str) -> str:
    raw = " ".join(str(label or "").split())
    if not raw:
        return raw
    key = raw.upper()
    if key in _ORG_ALIASES:
        return _ORG_ALIASES[key]
    return canonical_org(raw)


def _n(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return max(0, int(v))
    except (TypeError, ValueError):
        try:
            return max(0, int(float(v)))
        except (TypeError, ValueError):
            return 0


def parse_excel() -> list[dict]:
    ws = load_workbook(XLSX, data_only=True)["Arrival Status"]
    rows = []
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 2).value
        if not label:
            continue
        text = str(label).strip()
        low = text.lower()
        if low in ("grand total",) or low.startswith("legend") or low.startswith("source"):
            continue
        rows.append(
            {
                "excel_row": r,
                "label": text,
                "org": map_org(text),
                "planned": _n(ws.cell(r, 3).value),
                "arrived_strength": {
                    "gos_m": _n(ws.cell(r, 4).value),
                    "gos_f": _n(ws.cell(r, 5).value),
                    "sos_m": _n(ws.cell(r, 6).value),
                    "sos_f": _n(ws.cell(r, 7).value),
                    "ors_m": _n(ws.cell(r, 8).value),
                    "ors_f": _n(ws.cell(r, 9).value),
                },
                "arrived_total": _n(ws.cell(r, 10).value),
                "status": str(ws.cell(r, 13).value or "").strip(),
            }
        )
    return rows


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Missing workbook: {XLSX}")

    excel_rows = parse_excel()
    doc = store._read(store.FILES["arrival"], {"rows": [], "hubs": store.DEFAULT_HUBS})
    rows = doc.get("rows") or []
    hubs = store.normalize_hubs(doc.get("hubs"))

    # Drop bad row created by a mistaken API path segment.
    rows = [r for r in rows if not re.search(r"/arrived-strength$", r.get("org") or "", re.I)]

    by_org = {r["org"]: r for r in rows}
    by_org_key = {store._org_key(r["org"]): r for r in rows}

    updated = 0
    unmatched = []

    for x in excel_rows:
        org = x["org"]
        row = by_org.get(org) or by_org_key.get(store._org_key(org))
        if not row:
            canonical = store._resolve_org_name(org)
            row = by_org.get(canonical) or by_org_key.get(store._org_key(canonical))
        if not row:
            unmatched.append(x)
            continue

        cur = store.normalize_arrived_strength(row.get("arrived_strength"))
        cur.update(x["arrived_strength"])
        cur["updated_at"] = store._now()
        row["arrived_strength"] = cur
        updated += 1

    store._write(store.FILES["arrival"], {"hubs": hubs, "rows": rows})
    bundle = store.merge_bundle()

    print(f"Arrival status sync from {XLSX.name}")
    print(f"  Excel rows: {len(excel_rows)}")
    print(f"  Updated:    {updated}")
    if unmatched:
        print(f"  Unmatched:  {len(unmatched)}")
        for x in unmatched:
            print(f"    row {x['excel_row']}: {x['label']} -> {x['org']}")

    exp = arr = 0
    for u in bundle.get("units") or []:
        s = u.get("strength") or {}
        a = u.get("arrived_strength") or {}
        exp += sum(s.get(k) or 0 for k in store.STRENGTH_RANK_FIELDS)
        arr += sum(a.get(k) or 0 for k in store.STRENGTH_RANK_FIELDS)
    print(f"  Dashboard totals: {arr}/{exp} arrived")
    print("  Removed bad org row: Andhra Pradesh/arrived-strength (if present)")


if __name__ == "__main__":
    main()

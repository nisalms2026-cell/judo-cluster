"""
Import Total_Team_Final Excel into event_data.json for the ops dashboard.
"""
from __future__ import annotations

import json
import os
import re
import datetime
from openpyxl import load_workbook

from org_names import canonical_org

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, "event_data.json")


def find_latest_workbook():
    """Prefer newest Total_Team_Final*.xlsx by date in filename, then mtime."""
    import glob
    import re
    matches = glob.glob(os.path.join(BASE, "Total_Team_Final*.xlsx"))
    if not matches:
        return os.path.join(BASE, "Total_Team_Final 16.07.2026.xlsx")

    def sort_key(p):
        name = os.path.basename(p)
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", name)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return (y, mo, d, os.path.getmtime(p))
        return (0, 0, 0, os.path.getmtime(p))

    matches.sort(key=sort_key, reverse=True)
    return matches[0]


DEFAULT_XLSX = find_latest_workbook()


def _n(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default


def _s(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def _detect_location(cell_a):
    t = _s(cell_a).upper()
    if "TGPA" in t and t.startswith("1"):
        return "TGPA"
    if "NITHM" in t:
        return "NITHM"
    if "SPORT TOWER" in t:
        return "Sport Tower"
    if "OWN LOCATION" in t:
        return "Own Location"
    return None


def parse_final_and_mess(wb):
    """Build units from Final sheet; overlay mess from Mess Details."""
    ws = wb["Final"]
    location = "TGPA"
    units = {}
    order = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        loc_hit = _detect_location(a)
        if loc_hit:
            location = loc_hit
            continue
        org = canonical_org(_s(b).upper())
        if not org or org in ("ORGANISATION", "TOTAL", "GRAND TOTAL"):
            continue
        if "TOTAL" in org:
            continue
        # Must look like a data row (sl no numeric or blank with org)
        sl = a
        if sl is not None and not isinstance(sl, (int, float)) and not str(sl).isdigit():
            continue

        strength = {
            "gos_m": _n(ws.cell(r, 18).value),
            "sos_m": _n(ws.cell(r, 19).value),
            "ors_m": _n(ws.cell(r, 20).value),
            "gos_f": _n(ws.cell(r, 21).value),
            "sos_f": _n(ws.cell(r, 22).value),
            "ors_f": _n(ws.cell(r, 23).value),
        }
        total = _n(ws.cell(r, 24).value)
        if total == 0:
            total = sum(strength.values())

        unit = {
            "org": org,
            "location": location,
            "manager": {
                "name": _s(ws.cell(r, 3).value),
                "rank": _s(ws.cell(r, 4).value),
                "phone": _s(ws.cell(r, 5).value),
            },
            "count_gos": {"male": _n(ws.cell(r, 6).value), "female": _n(ws.cell(r, 7).value)},
            "count_sos": {"male": _n(ws.cell(r, 8).value), "female": _n(ws.cell(r, 9).value)},
            "support": {"male": _n(ws.cell(r, 10).value), "female": _n(ws.cell(r, 11).value)},
            "coach_male": {"sos": _n(ws.cell(r, 12).value), "ors": _n(ws.cell(r, 13).value)},
            "coach_female": {
                "gos": _n(ws.cell(r, 14).value),
                "sos": _n(ws.cell(r, 15).value),
                "ors": _n(ws.cell(r, 16).value),
            },
            "doctor": _n(ws.cell(r, 17).value),
            "strength": strength,
            "total": total,
            "mess": "",
            "travel": {"station": "", "arrival": "", "details": "", "status": "awaited"},
        }
        units[org] = unit
        order.append(org)

    # Mess Details overlay
    if "Mess Details" in wb.sheetnames:
        ws = wb["Mess Details"]
        location = "TGPA"
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            loc_hit = _detect_location(a)
            if loc_hit:
                location = loc_hit
                continue
            org = _s(ws.cell(r, 2).value).upper()
            if org not in units:
                continue
            mess = _s(ws.cell(r, 25).value)
            if mess:
                # Normalize spelling
                low = mess.lower()
                if "own" in low:
                    mess = "Own Arrangement"
                elif "tgpa" in low:
                    mess = "TGPA"
                elif "nihtm" in low or "nithm" in low:
                    mess = "NITHM"
                units[org]["mess"] = mess
            # Prefer mess-sheet manager if present (may be newer)
            name = _s(ws.cell(r, 3).value)
            if name:
                units[org]["manager"]["name"] = name
                units[org]["manager"]["rank"] = _s(ws.cell(r, 4).value) or units[org]["manager"]["rank"]
                units[org]["manager"]["phone"] = _s(ws.cell(r, 5).value) or units[org]["manager"]["phone"]

    return [units[o] for o in order if o in units]


def _is_mess_note_row(text: str) -> bool:
    """Excel sometimes puts ops notes in the org column (e.g. BSF stay at NITHM…)."""
    t = (text or "").upper()
    if not t:
        return False
    if "STAY AT" in t or "WILLING TO" in t or "DINE AT" in t:
        return True
    if "BSF" in t and "NITHM" in t and len(t) > 12:
        return True
    return False


def parse_tgpa_mess(wb):
    """TGPA Mess sheet: dining at TGPA vs own arrangement while staying at TGPA."""
    if "TGPA Mess" not in wb.sheetnames:
        return {"dining_tgpa": [], "own_mess": [], "note": ""}
    ws = wb["TGPA Mess"]
    dining = []
    own = []
    note = ""

    for r in range(3, ws.max_row + 1):
        left_org = _s(ws.cell(r, 2).value)
        if left_org and left_org.upper() != "TOTAL":
            if _is_mess_note_row(left_org):
                note = left_org
            else:
                dining.append({
                    "org": canonical_org(left_org),
                    "iom": _n(ws.cell(r, 3).value),
                    "jom1": _n(ws.cell(r, 4).value),
                    "jom2": _n(ws.cell(r, 5).value),
                    "sotc": _n(ws.cell(r, 6).value),
                    "total": _n(ws.cell(r, 7).value),
                })
        right_org = _s(ws.cell(r, 10).value)
        if right_org and right_org.upper() != "TOTAL":
            if _is_mess_note_row(right_org):
                note = right_org
            else:
                own.append({
                    "org": canonical_org(right_org),
                    "iom": _n(ws.cell(r, 11).value),
                    "jom1": _n(ws.cell(r, 12).value),
                    "jom2": _n(ws.cell(r, 13).value),
                    "sotc": _n(ws.cell(r, 14).value),
                    "total": _n(ws.cell(r, 15).value),
                })
        # Notes in other columns
        for c in range(1, 16):
            v = _s(ws.cell(r, c).value)
            if _is_mess_note_row(v):
                note = v

    return {"dining_tgpa": dining, "own_mess": own, "note": note}


def _travel_sheet(wb):
    """Prefer dedicated arrival-plan sheet names, then classic Travel Details."""
    for name in ("Travel Details Final", "Travel Details", "Arrival Plan"):
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        low = name.lower()
        if "travel" in low or "arrival" in low:
            return wb[name]
    return None


def _split_multiline(val: str):
    """Split Excel cells that pack multiple legs with newlines."""
    parts = [p.strip() for p in re.split(r"[\r\n]+", _s(val)) if p.strip()]
    return parts


def parse_travel(wb, units_by_org):
    ws = _travel_sheet(wb)
    if ws is None:
        return
    station = ""
    in_awaited_block = False
    in_flight_block = False
    in_bus_block = False
    for r in range(2, ws.max_row + 1):
        a = _s(ws.cell(r, 1).value)
        b = _s(ws.cell(r, 2).value)
        c3 = _s(ws.cell(r, 3).value)
        c4 = _s(ws.cell(r, 4).value)

        # Section headers
        header_blob = f"{c3} {c4}".upper()
        if "TRAVEL PLAN AWAITED" in header_blob:
            in_awaited_block = True
            in_flight_block = False
            in_bus_block = False
            station = ""
            continue
        if "ARRIVAL BY FLIGHT" in header_blob:
            in_awaited_block = False
            in_flight_block = True
            in_bus_block = False
            station = "Rajiv Gandhi Intl Airport (Shamshabad)"
            continue
        if "TRAVEL BY BUS" in header_blob or (c4.upper() == "TRAVEL BY BUS"):
            in_awaited_block = False
            in_flight_block = False
            in_bus_block = True
            station = "Bus Alighting Point"
            continue

        # Station name in col B when present
        if b and any(k in b.upper() for k in ("JN", "STATION", "RLY", "HYDERABAD", "SECUNDERABAD", "CHARLAPALLI", "KACHEGUDA", "AIRPORT", "FLIGHT")):
            station = b
            in_awaited_block = False
            in_flight_block = False
            in_bus_block = False
        if a and any(k in a.upper() for k in ("SECUNDERABAD", "HYDERABAD", "CHARLAPALLI", "KACHEGUDA")):
            station = a if not b else b
            in_awaited_block = False
            in_flight_block = False
            in_bus_block = False
        if a.upper() == "BUS" or b.upper() == "BUS" or "ARRIVAL BY BUS" in header_blob:
            station = "Bus Alighting Point"
            in_bus_block = True
            in_awaited_block = False
            in_flight_block = False
            if not c4:
                continue

        org_raw = c4
        if not org_raw or org_raw.upper() in ("TEAM NAME", "ARRIVAL BY FLIGHT", "TRAVEL BY BUS", "TRAVEL PLAN AWAITED"):
            continue

        org_key = _match_org(org_raw, units_by_org)
        arrival_raw = _s(ws.cell(r, 7).value)
        details_raw = _s(ws.cell(r, 8).value)
        mgr = _s(ws.cell(r, 5).value)
        phone = _s(ws.cell(r, 6).value)

        arrivals = _split_multiline(arrival_raw)
        details_parts = _split_multiline(details_raw)
        # If one details blob for many times, keep full text on primary and extras share it
        if len(details_parts) <= 1 and details_raw:
            details_parts = [details_raw]

        is_bus = (
            in_bus_block
            or station.upper() in ("BUS", "BUS ALIGHTING POINT")
            or "BY BUS" in details_raw.upper()
            or details_raw.upper() == "BUS"
        )
        is_flight = (
            in_flight_block
            or "FLIGHT" in (station + " " + details_raw).upper()
            or "AIRPORT" in (station + " " + details_raw).upper()
            or "AIR INDIA" in details_raw.upper()
            or "INDIGO" in details_raw.upper()
            or "AKASA" in details_raw.upper()
        )

        row_station = "Bus Alighting Point" if is_bus else (
            "Rajiv Gandhi Intl Airport (Shamshabad)" if is_flight and (not station or "FLIGHT" in station.upper() or "AIRPORT" in station.upper())
            else (station or _s(ws.cell(r, 2).value))
        )

        # Build one entry per arrival time when multiple are listed
        n_legs = max(len(arrivals), 1)
        for i in range(n_legs):
            arrival = arrivals[i] if i < len(arrivals) else (arrivals[0] if arrivals else "")
            if i < len(details_parts):
                details = details_parts[i]
            elif details_raw:
                details = details_raw
            else:
                details = "By Bus" if is_bus else ""

            if in_awaited_block:
                status = "awaited"
            elif arrival or details:
                # Includes "By Bus" — plan mode known, ETA may still be blank
                status = "planned"
            else:
                status = "awaited"

            entry = {
                "mode": "bus" if is_bus else ("flight" if is_flight else "rail"),
                "station": row_station,
                "arrival": arrival,
                "details": details or ("By Bus" if is_bus else ""),
                "status": status,
                "team_label": org_raw.replace("\n", " ").strip(),
            }
            if is_bus:
                entry["bus_role"] = "alighting"
            elif entry["mode"] == "flight" and (not entry["station"] or entry["station"].upper() in ("FLIGHT / AIRPORT", "FLIGHT", "AIRPORT")):
                entry["station"] = "Rajiv Gandhi Intl Airport (Shamshabad)"

            if org_key and org_key in units_by_org:
                u = units_by_org[org_key]
                existing = u.get("travel") or {}
                if i == 0:
                    if existing.get("status") != "planned" and not existing.get("arrival") and not existing.get("details"):
                        u["travel"] = entry
                    elif status == "planned" and existing.get("status") != "planned":
                        u["travel"] = entry
                    elif status == "planned" and existing.get("status") == "planned":
                        # Another planned row for same org (e.g. ITBP / Himachal half) → extra
                        if existing.get("arrival") != entry.get("arrival") or existing.get("details") != entry.get("details") or existing.get("station") != entry.get("station"):
                            u.setdefault("travel_extra", []).append(entry)
                        else:
                            u["travel"] = entry
                    else:
                        if existing.get("status") != "planned":
                            u["travel"] = entry
                else:
                    u.setdefault("travel_extra", []).append(entry)

                if mgr:
                    u["manager"]["name"] = mgr
                if phone:
                    u["manager"]["phone"] = str(phone)
            elif org_key is None:
                # Keep unmatched labels visible for ops review via orphan bucket
                orphans = units_by_org.setdefault("__travel_orphans__", {"org": "__travel_orphans__", "travel": {}, "travel_extra": [], "manager": {}})
                orphans.setdefault("travel_extra", []).append({**entry, "team_label": org_raw})


def _match_org(raw, units_by_org):
    raw_u = raw.upper().strip()
    # Strip parenthetical notes
    raw_u = re.sub(r"\(.*?\)", "", raw_u).strip()
    raw_u = re.sub(r"\s+", " ", raw_u)

    aliases = {
        "ASSAM RIFLES": "ASSAM RIFLES",
        "HIMACHAL PRADESH POLICE": "HIMACHAL PRADESH POLICE",
        "BORDER SECURITY FORCE": "BSF",
        "BSF": "BSF",
        "ODISHA": "ORISSA POLICE",
        "ORISSA": "ORISSA POLICE",
        "ITBP": "ITBP",
        "ITBP WUSHU TEAM": "ITBP",
        "SSB": "SSB",
        "HARYANA POLICE": "HARYANA",
        "HARYANA": "HARYANA",
        "KERALA POLICE": "KERALA POLICE",
        "KARNATAKA STATE POLICE": "KARNATAKA POLICE",
        "KARNATAKA POLICE": "KARNATAKA POLICE",
        "CHHATTISGARH": "CHATTISGARH",
        "CHATTISGARH": "CHATTISGARH",
        "UTTARAKHAND": "UTTARAKHAND POLICE",
        "GUJRAT": canonical_org("Gujrat"),
        "GUJARAT": canonical_org("Gujrat"),
        "BIHAR": "BIHAR POLICE",
        "UTTARPRADESH": "UTTARPRADESH",
        "J&K POLICE": canonical_org("JAMMU & KASHMIR"),
        "J&K": canonical_org("JAMMU & KASHMIR"),
        "JAMMU & KASHMIR": canonical_org("JAMMU & KASHMIR"),
        "PUNJAB": "PUNJAB POLICE",
        "MAHARASHTRA POLICE": "MAHARASHTRA POLICE",
        "RAJASTHAN": "RAJASTHAN POLICE",
        "TAMIL NADU": "TAMILNADU",
        "TAMILNADU": "TAMILNADU",
        "WEST BENGAL POLICE": "WEST BENGAL POLICE",
        "RPF": "RPF",
        "ASSAM POLICE": "ASSAM POLICE",
        "MADHYA PRADESH": "MADHYA PRADESH POLICE",
        "ANDHRA POLICE": canonical_org("Andhra Pradesh"),
        "ANDHRA PRADESH POLICE": canonical_org("Andhra Pradesh"),
        "ANDHRA PRADESH": canonical_org("Andhra Pradesh"),
        "ARUNACHAL PRADESH": "ARUNACHAL PRADESH",
        "CHANDIGARH POLICE": "CHANDIGARH POLICE",
        "DELHI POLICE": "DELHI POLICE",
        "GOA": "GOA",
        "NSG": "NSG",
        "JHARKHAND": "JHARKHAND",
        "CISF": "CISF",
        "TELANGANA POLICE": "TELANGANA POLICE",
        "CRPF": "CRPF",
        "MIZORAM POLICE": "MIZORAM POLICE",
        "MANIPUR POLICE": "MANIPUR POLICE",
        "SIKKIM": "SIKKIM POLICE",
        "SIKKIM POLICE": "SIKKIM POLICE",
    }
    if raw_u in aliases:
        return aliases[raw_u]
    for k, v in aliases.items():
        if k in raw_u or raw_u in k:
            return v
    for org in units_by_org:
        if org == "__travel_orphans__":
            continue
        if org in raw_u or raw_u in org:
            return org
    return None


def mark_awaited_from_travel_sheet(wb, units_by_org):
    """Force awaited for orgs listed under 'Travel plan awaited'."""
    ws = _travel_sheet(wb)
    if ws is None:
        return
    in_awaited = False
    for r in range(1, ws.max_row + 1):
        c3 = _s(ws.cell(r, 3).value).upper()
        c4 = _s(ws.cell(r, 4).value)
        blob = f"{c3} {c4.upper()}"
        if "TRAVEL PLAN AWAITED" in blob:
            in_awaited = True
            continue
        if ("ARRIVAL BY FLIGHT" in blob or "TRAVEL BY BUS" in blob) and not c4:
            in_awaited = False
            continue
        if in_awaited and c4:
            key = _match_org(c4, units_by_org)
            if key and key in units_by_org:
                u = units_by_org[key]
                note = _s(ws.cell(r, 8).value)
                mgr = _s(ws.cell(r, 5).value)
                phone = _s(ws.cell(r, 6).value)
                u["travel"] = {
                    "station": "",
                    "arrival": "",
                    "details": note,
                    "status": "awaited",
                    "team_label": c4,
                }
                if mgr:
                    u["manager"]["name"] = mgr
                if phone:
                    u["manager"]["phone"] = str(phone)
                u.pop("travel_extra", None)


def import_arrival_plan(xlsx_path=None):
    """Import only travel/arrival from Arrival Plan workbook into arrival.json."""
    from store import (
        FILES,
        _read,
        _write,
        normalize_hubs,
        merge_bundle,
    )

    path = xlsx_path or os.path.join(BASE, "Arrival Plan 17.07.2026.xlsx")
    if not os.path.isfile(path):
        # fallback: newest Arrival Plan*.xlsx
        import glob
        matches = glob.glob(os.path.join(BASE, "Arrival Plan*.xlsx"))
        if not matches:
            raise FileNotFoundError(f"Arrival plan Excel not found: {path}")
        matches.sort(key=os.path.getmtime, reverse=True)
        path = matches[0]

    acc = _read(FILES["accommodation"], {"rows": []})
    existing = _read(FILES["arrival"], {"rows": [], "hubs": {}})
    dep_map = {
        r["org"]: r.get("travel_departure")
        for r in (existing.get("rows") or [])
        if r.get("travel_departure")
    }

    by_org = {}
    for u in acc.get("rows") or []:
        org = u.get("org")
        if not org:
            continue
        by_org[org] = {
            "org": org,
            "manager": {"name": "", "rank": "", "phone": ""},
            "travel": {"station": "", "arrival": "", "details": "", "status": "awaited"},
            "travel_extra": [],
        }

    wb = load_workbook(path, data_only=True)
    parse_travel(wb, by_org)
    mark_awaited_from_travel_sheet(wb, by_org)

    rows = []
    for org, u in by_org.items():
        if org == "__travel_orphans__":
            continue
        row = {
            "org": org,
            "travel": u.get("travel") or {"station": "", "arrival": "", "details": "", "status": "awaited"},
            "travel_extra": u.get("travel_extra") or [],
        }
        if dep_map.get(org):
            row["travel_departure"] = dep_map[org]
        rows.append(row)

    # Keep accommodation order
    order = [u["org"] for u in (acc.get("rows") or []) if u.get("org")]
    rank = {o: i for i, o in enumerate(order)}
    rows.sort(key=lambda r: rank.get(r["org"], 999))

    _write(FILES["arrival"], {
        "hubs": normalize_hubs(existing.get("hubs")),
        "rows": rows,
        "source_file": os.path.basename(path),
        "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
    })

    # Also refresh directory phones/names from travel sheet managers when present
    directory = _read(FILES["directory"], {"rows": []})
    dir_map = {r["org"]: r for r in (directory.get("rows") or [])}
    for org, u in by_org.items():
        if org == "__travel_orphans__":
            continue
        mgr = u.get("manager") or {}
        if not (mgr.get("name") or mgr.get("phone")):
            continue
        cur = dir_map.get(org) or {"org": org, "manager": {"name": "", "rank": "", "phone": ""}}
        m = cur.setdefault("manager", {"name": "", "rank": "", "phone": ""})
        if mgr.get("name"):
            m["name"] = mgr["name"]
        if mgr.get("phone"):
            m["phone"] = str(mgr["phone"])
        dir_map[org] = cur
    _write(FILES["directory"], {
        "rows": [dir_map[o] for o in order if o in dir_map] + [r for o, r in dir_map.items() if o not in rank],
        "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
    })

    return merge_bundle()


def build_summary(units):
    locs = {}
    mess = {}
    travel_planned = 0
    travel_awaited = 0
    g = {"gos_m": 0, "sos_m": 0, "ors_m": 0, "gos_f": 0, "sos_f": 0, "ors_f": 0}
    for u in units:
        loc = u["location"]
        locs.setdefault(loc, {"teams": 0, "strength": 0})
        locs[loc]["teams"] += 1
        locs[loc]["strength"] += u["total"]
        m = u.get("mess") or "Unassigned"
        mess[m] = mess.get(m, 0) + u["total"]
        for k in g:
            g[k] += u["strength"].get(k, 0)
        st = u.get("travel", {}).get("status", "awaited")
        if st == "planned":
            travel_planned += 1
        else:
            travel_awaited += 1
    male = g["gos_m"] + g["sos_m"] + g["ors_m"]
    female = g["gos_f"] + g["sos_f"] + g["ors_f"]
    return {
        "teams": len(units),
        "strength": sum(u["total"] for u in units),
        "male": male,
        "female": female,
        "gos": g["gos_m"] + g["gos_f"],
        "sos": g["sos_m"] + g["sos_f"],
        "ors": g["ors_m"] + g["ors_f"],
        "by_location": locs,
        "by_mess": mess,
        "travel_planned": travel_planned,
        "travel_awaited": travel_awaited,
        "strength_detail": g,
    }


def import_workbook(xlsx_path=None):
    xlsx_path = xlsx_path or find_latest_workbook()
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Excel not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    units = parse_final_and_mess(wb)
    by_org = {u["org"]: u for u in units}
    parse_travel(wb, by_org)
    mark_awaited_from_travel_sheet(wb, by_org)
    # refresh list from dict (travel mutations)
    units = [by_org[u["org"]] for u in units]
    tgpa_mess = parse_tgpa_mess(wb)

    data = {
        "updated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source_file": os.path.basename(xlsx_path),
        "event": {
            "title": "11th All India Police Judo Cluster 2026",
            "host": "CISF",
            "city": "Hyderabad",
            "mascot": "Vira",
        },
        "units": units,
        "venues": sorted({u.get("location") for u in units if u.get("location")}),
        "tgpa_mess": tgpa_mess,
        "summary": build_summary(units),
    }
    # Lazy import avoids circular dependency at module load
    from store import save_imported_bundle
    return save_imported_bundle(data)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] in ("--arrival", "arrival"):
        path = sys.argv[2] if len(sys.argv) > 2 else None
        d = import_arrival_plan(path)
        print(f"Arrival plan imported -> data/arrival.json")
        print(f"Travel planned {d['summary']['travel_planned']} / awaited {d['summary']['travel_awaited']}")
    else:
        d = import_workbook()
        print(f"Imported {len(d['units'])} units -> data/*.json")
        print(f"Strength {d['summary']['strength']} | Travel planned {d['summary']['travel_planned']} / awaited {d['summary']['travel_awaited']}")

"""
Import player roster from AIPSCB MIS Report Excel into data/players.json.
"""
from __future__ import annotations

import glob
import os
import re
import uuid
from openpyxl import load_workbook

from org_names import ANDHRA_PRADESH, GUJRAT, JAMMU_KASHMIR, canonical_org
BASE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_MIS = os.path.join(BASE, "AIPSCB - MIS Report.xlsx")

PLAYER_SPORTS = ["Judo", "Karate", "Taekwondo", "Wushu", "Pencak Silat", "Taolu"]

SPORT_SPLIT_RE = re.compile(
    r"(?=(?:Pencak Silat|Taolu|Taekwondo|Karate|Wushu|Judo)\s*\()",
    re.IGNORECASE,
)

ORG_ALIASES = {
    "ANDHRA PRADESH": ANDHRA_PRADESH,
    "ARUNACHAL PRADESH": "ARUNACHAL PRADESH",
    "ASSAM": "ASSAM POLICE",
    "ASSAM RIFLES": "ASSAM RIFLES",
    "BIHAR": "BIHAR POLICE",
    "BSF": "BSF",
    "CHANDIGARH": "CHANDIGARH POLICE",
    "CHHATTISGARH": "CHHATTISGARH POLICE",
    "CISF": "CISF",
    "CRPF": "CRPF",
    "DELHI": "DELHI POLICE",
    "GOA": "GOA POLICE",
    "GUJARAT": GUJRAT,
    "HARYANA": "HARYANA POLICE",
    "HIMACHAL PRADESH": "HIMACHAL PRADESH POLICE",
    "ITBP": "ITBP",
    "JAMMU AND KASHMIR": JAMMU_KASHMIR,
    "J&K": JAMMU_KASHMIR,
    "JHARKHAND": "JHARKHAND POLICE",
    "KARNATAKA": "KARNATAKA POLICE",
    "KERALA": "KERALA POLICE",
    "LADAKH": "LADAKH POLICE",
    "MADHYA PRADESH": "MADHYA PRADESH POLICE",
    "MAHARASHTRA": "MAHARASHTRA POLICE",
    "MANIPUR": "MANIPUR POLICE",
    "MEGHALAYA": "MEGHALAYA POLICE",
    "MIZORAM": "MIZORAM POLICE",
    "NAGALAND": "NAGALAND POLICE",
    "NSG": "NSG",
    "ODISHA": "ODISHA POLICE",
    "ORISSA": "ODISHA POLICE",
    "PUNJAB": "PUNJAB POLICE",
    "RAJASTHAN": "RAJASTHAN POLICE",
    "RPF": "RPF",
    "SIKKIM": "SIKKIM POLICE",
    "SSB": "SSB",
    "TAMIL NADU": "TAMIL NADU POLICE",
    "TAMILNADU": "TAMIL NADU POLICE",
    "TELANGANA": "TELANGANA POLICE",
    "TRIPURA": "TRIPURA POLICE",
    "UTTAR PRADESH": "UTTAR PRADESH POLICE",
    "UTTARAKHAND": "UTTARAKHAND POLICE",
    "WEST BENGAL": "WEST BENGAL POLICE",
}


def find_mis_workbook():
    matches = glob.glob(os.path.join(BASE, "AIPSCB - MIS Report*.xlsx"))
    if not matches:
        return DEFAULT_MIS
    matches.sort(key=os.path.getmtime, reverse=True)
    return matches[0]


def _s(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def _digits(v):
    return re.sub(r"\D", "", _s(v))


def _n(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return default


def normalize_org(raw: str) -> tuple[str, str]:
    raw = _s(raw)
    key = re.sub(r"\s+", " ", raw.upper())
    key = key.replace("POLICE POLICE", "POLICE")
    mapped = ORG_ALIASES.get(key)
    if mapped:
        return canonical_org(mapped), raw
    if key.endswith(" POLICE"):
        return canonical_org(key), raw
    if key in ("CISF", "CRPF", "BSF", "ITBP", "SSB", "NSG", "RPF", "ASSAM RIFLES"):
        return canonical_org(key), raw
    guess = ORG_ALIASES.get(key.replace(" POLICE", ""))
    if guess:
        return canonical_org(guess), raw
    return canonical_org(key or raw.upper()), raw


def _detect_sport(label: str) -> str:
    t = label.lower()
    for sport in PLAYER_SPORTS:
        if sport.lower() in t:
            return sport
    return "Unknown"


def parse_events(text: str) -> list[dict]:
    text = _s(text)
    if not text:
        return []
    parts = [p.strip() for p in SPORT_SPLIT_RE.split(text) if p.strip()]
    if not parts:
        return [{"sport": _detect_sport(text), "label": text}]
    out = []
    for part in parts:
        sport = _detect_sport(part)
        out.append({"sport": sport, "label": part})
    return out


def parse_qualified(text: str) -> list[str]:
    text = _s(text)
    if not text:
        return []
    return parse_events(text)  # same format when populated


def _player_id(mobile: str, name: str, org: str, row_no: int) -> str:
    if len(mobile) == 10:
        return f"pl_{mobile}"
    slug = re.sub(r"[^a-z0-9]+", "", f"{name}{org}".lower())[:24]
    return f"pl_{slug}_{row_no}"


def import_mis_report(xlsx_path=None) -> dict:
    path = xlsx_path or find_mis_workbook()
    if not os.path.isfile(path):
        raise FileNotFoundError(path)

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    players = []
    seen_ids = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = list(row)
        if header is None:
            if any(_s(v) == "Name" for v in vals):
                header = [_s(v) for v in vals]
            continue
        if not header:
            continue
        d = {header[j]: vals[j] if j < len(vals) else None for j in range(len(header))}
        name = _s(d.get("Name"))
        if not name:
            continue
        org, org_raw = normalize_org(d.get("Organization Name"))
        mobile = _digits(d.get("Contact No.") or d.get("Contact No"))
        email = _s(d.get("Email")).lower()
        gender = _s(d.get("Gender"))
        participated = _s(d.get("TournamentParticipated"))
        qualified_raw = _s(d.get("TournamentQualified"))
        events = parse_events(participated)
        qualified = [q["label"] for q in parse_qualified(qualified_raw)] if qualified_raw else []

        pid = _player_id(mobile, name, org, i + 1)
        if pid in seen_ids:
            pid = f"{pid}_{uuid.uuid4().hex[:6]}"
        seen_ids.add(pid)

        players.append({
            "id": pid,
            "sno": _n(d.get("Sr.no")),
            "name": name,
            "org": org,
            "org_raw": org_raw,
            "gender": gender,
            "mobile": mobile,
            "email": email,
            "events": events,
            "qualified": qualified,
        })

    wb.close()

    payload = {
        "source_file": os.path.basename(path),
        "sports": list(PLAYER_SPORTS),
        "players": players,
    }
    return payload


def import_and_save(xlsx_path=None) -> dict:
    import store

    payload = import_mis_report(xlsx_path)
    return store.save_players_import(payload)


def main():
    import argparse

    ap = argparse.ArgumentParser(description="Import AIPSCB MIS player roster")
    ap.add_argument("xlsx", nargs="?", help="Path to MIS Report xlsx")
    args = ap.parse_args()
    bundle = import_and_save(args.xlsx)
    players = (bundle.get("players") or {}).get("players") or []
    print(f"Imported {len(players)} players from MIS Report")
    by_sport = {}
    for p in players:
        for ev in p.get("events") or []:
            sp = ev.get("sport") or "Unknown"
            by_sport[sp] = by_sport.get(sp, 0) + 1
    for sp in PLAYER_SPORTS:
        print(f"  {sp}: {by_sport.get(sp, 0)} event entries")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
